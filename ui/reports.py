import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QComboBox, QFrame, QMessageBox, QFileDialog, QGridLayout)
from PySide6.QtCore import Qt
from datetime import datetime

BTN_STYLE = """
    QPushButton {
        background: qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #0d1832,stop:1 #0a1428);
        border: 1px solid #1e3a5f; border-radius: 10px;
        color: #c0d0e0; font-size: 13px; padding: 20px 15px;
        text-align: center;
    }
    QPushButton:hover { border-color: #00d4ff; color: #00d4ff; background: rgba(0,212,255,0.08); }
    QPushButton:pressed { background: rgba(0,212,255,0.15); }
"""


class ReportsWidget(QWidget):
    def __init__(self, session):
        super().__init__()
        self.session = session
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self); layout.setContentsMargins(20,20,20,20); layout.setSpacing(20)

        title = QLabel('📊  Rapports & Exports')
        title.setStyleSheet('color: #00d4ff; font-size: 18px; font-weight: bold; background: transparent;')
        layout.addWidget(title)

        # Year filter
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel('Année scolaire:'))
        self.year_combo = QComboBox()
        year = datetime.now().year
        self.year_combo.addItems([str(y) for y in range(year, year-5, -1)])
        self.year_combo.setStyleSheet('background: #0d1832; border: 1px solid #1e3a5f; border-radius: 6px; color: #e0e6f0; padding: 6px 10px; min-width: 100px;')
        filter_row.addWidget(self.year_combo)
        filter_row.addStretch()
        layout.addLayout(filter_row)

        grid = QGridLayout(); grid.setSpacing(14)

        reports = [
            ('📋  Liste des Élèves\nExport PDF complet', self._export_students_pdf, '#00d4ff'),
            ('💰  Rapport Financier\nRevenus & Dépenses annuels', self._export_financial_pdf, '#00ff88'),
            ('💳  Historique Paiements\nExport Excel (.xlsx)', self._export_payments_excel, '#8866ff'),
            ('👥  Liste du Personnel\nEnseignants & Employés', self._export_employees_pdf, '#ffaa00'),
            ('📅  Emploi du Temps\nPDF par classe', self._export_timetable_pdf, '#ff88aa'),
            ('🚌  Rapport Transport\nAbonnés & Bus', self._export_transport_pdf, '#00aaff'),
            ('📊  Analyse Paiements\nRépartition mensuelle Excel', self._export_analytics_excel, '#ff6644'),
            ('💾  Sauvegarde DB\nExport base de données', self._backup_db, '#aaffaa'),
        ]

        for idx, (label, func, color) in enumerate(reports):
            btn = QPushButton(label)
            btn.setStyleSheet(BTN_STYLE + f'QPushButton:hover {{ border-color: {color}; color: {color}; }}')
            btn.setMinimumHeight(80)
            btn.clicked.connect(func)
            grid.addWidget(btn, idx // 2, idx % 2)

        layout.addLayout(grid)
        layout.addStretch()

        self.status_lbl = QLabel()
        self.status_lbl.setStyleSheet('color: #00ff88; font-size: 12px; background: transparent;')
        layout.addWidget(self.status_lbl)

    def _export_students_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import mm
            from models.database import Student

            path, _ = QFileDialog.getSaveFileName(self, 'Enregistrer', f'eleves_{datetime.now().year}.pdf', 'PDF (*.pdf)')
            if not path: return

            students = self.session.query(Student).filter_by(active=True).order_by(Student.class_name, Student.last_name).all()
            doc = SimpleDocTemplate(path, pagesize=A4, rightMargin=15*mm, leftMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
            styles = getSampleStyleSheet()
            story = []

            title_style = ParagraphStyle('t', fontSize=18, textColor=colors.HexColor('#0044cc'),
                                         alignment=1, fontName='Helvetica-Bold', spaceAfter=6)
            story.append(Paragraph('Le Schéma — Liste des Élèves', title_style))
            story.append(Paragraph(f'Généré le {datetime.now().strftime("%d/%m/%Y")} | Total: {len(students)} élèves', styles['Normal']))
            story.append(Spacer(1, 8*mm))

            data = [['Code', 'Nom', 'Prénom', 'Classe', 'Parent', 'Téléphone', 'Transport']]
            for s in students:
                data.append([s.code or '', s.last_name or '', s.first_name or '',
                              s.class_name or '', s.parent_name or '',
                              s.parent_phone or '', '✓' if s.transport else ''])

            t = Table(data, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0044cc')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f0f4ff')]),
                ('PADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(t)
            doc.build(story)
            self.status_lbl.setText(f'✅ Export élèves: {path}')
        except Exception as e:
            QMessageBox.critical(self, 'Erreur', str(e))

    def _export_financial_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import mm
            from models.database import Payment, Expense
            from themes.style import MONTHS

            path, _ = QFileDialog.getSaveFileName(self, 'Enregistrer', f'rapport_financier_{datetime.now().year}.pdf', 'PDF (*.pdf)')
            if not path: return

            year = int(self.year_combo.currentText())
            payments = self.session.query(Payment).filter(Payment.year == year).all()
            expenses = self.session.query(Expense).filter(Expense.date != None).all()
            expenses_year = [e for e in expenses if e.date and e.date.year == year]

            total_rev = sum(p.amount for p in payments)
            total_exp = sum(e.amount for e in expenses_year)
            profit = total_rev - total_exp

            doc = SimpleDocTemplate(path, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm,
                                    topMargin=15*mm, bottomMargin=15*mm)
            styles = getSampleStyleSheet()
            story = []

            title_style = ParagraphStyle('t', fontSize=20, textColor=colors.HexColor('#0044cc'),
                                          alignment=1, fontName='Helvetica-Bold', spaceAfter=6)
            story.append(Paragraph(f'Rapport Financier {year} — Le Schéma', title_style))
            story.append(Paragraph(f'Généré le {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
            story.append(Spacer(1, 8*mm))
            story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#0044cc')))
            story.append(Spacer(1, 5*mm))

            summary = [
                ['REVENUS TOTAUX', f'{total_rev:.2f} MAD'],
                ['DÉPENSES TOTALES', f'{total_exp:.2f} MAD'],
                ['BÉNÉFICE NET', f'{profit:.2f} MAD'],
            ]
            st = Table(summary, colWidths=[100*mm, 60*mm])
            st.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e8f0ff')),
                ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#ffe8e8')),
                ('BACKGROUND', (0,2), (-1,2), colors.HexColor('#e8ffe8') if profit >= 0 else colors.HexColor('#ffe8e8')),
                ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 12),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('PADDING', (0,0), (-1,-1), 10),
                ('ALIGN', (1,0), (1,-1), 'RIGHT'),
            ]))
            story.append(st)
            story.append(Spacer(1, 8*mm))

            # Monthly breakdown
            story.append(Paragraph('Revenus par mois', ParagraphStyle('sh', fontSize=13,
                textColor=colors.HexColor('#0044cc'), fontName='Helvetica-Bold', spaceAfter=4)))
            month_data = [['Mois', 'Mensualités', 'Assurance', 'Transport', 'Total']]
            for m in MONTHS:
                monthly = sum(p.amount for p in payments if p.month == m and p.payment_type == 'monthly')
                insurance = sum(p.amount for p in payments if p.month == m and p.payment_type == 'insurance')
                transport = sum(p.amount for p in payments if p.month == m and p.payment_type == 'transport')
                total = monthly + insurance + transport
                if total > 0:
                    month_data.append([m, f'{monthly:.0f}', f'{insurance:.0f}', f'{transport:.0f}', f'{total:.0f} MAD'])
            mt = Table(month_data, repeatRows=1)
            mt.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0044cc')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f0f4ff')]),
                ('PADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(mt)
            doc.build(story)
            self.status_lbl.setText(f'✅ Rapport financier: {path}')
        except Exception as e:
            QMessageBox.critical(self, 'Erreur', str(e))

    def _export_payments_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from models.database import Payment, Student

            path, _ = QFileDialog.getSaveFileName(self, 'Enregistrer', f'paiements_{datetime.now().year}.xlsx', 'Excel (*.xlsx)')
            if not path: return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Paiements'

            header_fill = PatternFill(start_color='0044CC', end_color='0044CC', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            headers = ['N° Reçu','Élève','Classe','Type','Mois','Année','Montant (MAD)','Date']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill; cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            payments = self.session.query(Payment).order_by(Payment.payment_date.desc()).all()
            for row, p in enumerate(payments, 2):
                student = self.session.query(Student).filter_by(id=p.student_id).first()
                name = f'{student.first_name} {student.last_name}' if student else '-'
                cls = student.class_name if student else '-'
                ws.cell(row=row, column=1, value=p.receipt_number or '')
                ws.cell(row=row, column=2, value=name)
                ws.cell(row=row, column=3, value=cls)
                ws.cell(row=row, column=4, value=p.payment_type or '')
                ws.cell(row=row, column=5, value=p.month or '')
                ws.cell(row=row, column=6, value=p.year or '')
                ws.cell(row=row, column=7, value=p.amount or 0)
                ws.cell(row=row, column=8, value=str(p.payment_date)[:19] if p.payment_date else '')
                if row % 2 == 0:
                    for col in range(1, 9):
                        ws.cell(row=row, column=col).fill = PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col) + 3
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 30)

            wb.save(path)
            self.status_lbl.setText(f'✅ Export Excel: {path}')
        except Exception as e:
            QMessageBox.critical(self, 'Erreur', str(e))

    def _export_employees_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import mm
            from models.database import Employee

            path, _ = QFileDialog.getSaveFileName(self, 'Enregistrer', f'personnel_{datetime.now().year}.pdf', 'PDF (*.pdf)')
            if not path: return

            employees = self.session.query(Employee).filter_by(active=True).order_by(Employee.role).all()
            doc = SimpleDocTemplate(path, pagesize=A4, rightMargin=15*mm, leftMargin=15*mm,
                                    topMargin=15*mm, bottomMargin=15*mm)
            styles = getSampleStyleSheet()
            story = []

            title_style = ParagraphStyle('t', fontSize=16, textColor=colors.HexColor('#0044cc'),
                                          alignment=1, fontName='Helvetica-Bold', spaceAfter=6)
            story.append(Paragraph('Le Schéma — Liste du Personnel', title_style))
            story.append(Paragraph(f'Total: {len(employees)} personnes', styles['Normal']))
            story.append(Spacer(1, 8*mm))

            data = [['Nom', 'Prénom', 'Poste', 'Téléphone', 'Email', 'Date embauche', 'Salaire']]
            for e in employees:
                data.append([e.last_name or '', e.first_name or '', e.role or '',
                              e.phone or '', e.email or '', str(e.hire_date) if e.hire_date else '',
                              f'{e.base_salary:.0f} MAD'])
            t = Table(data, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0044cc')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f0f4ff')]),
                ('PADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(t)
            doc.build(story)
            self.status_lbl.setText(f'✅ Export personnel: {path}')
        except Exception as e:
            QMessageBox.critical(self, 'Erreur', str(e))

    def _export_timetable_pdf(self):
        self.status_lbl.setText('ℹ️ Export emploi du temps: utilisez le module Emploi du Temps.')

    def _export_transport_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import mm
            from models.database import Student, Bus, Employee

            path, _ = QFileDialog.getSaveFileName(self, 'Enregistrer', f'transport_{datetime.now().year}.pdf', 'PDF (*.pdf)')
            if not path: return

            students = self.session.query(Student).filter_by(active=True, transport=True).all()
            buses = self.session.query(Bus).filter_by(active=True).all()
            doc = SimpleDocTemplate(path, pagesize=A4, rightMargin=15*mm, leftMargin=15*mm,
                                    topMargin=15*mm, bottomMargin=15*mm)
            story = []
            title_style = ParagraphStyle('t', fontSize=16, textColor=colors.HexColor('#0044cc'),
                                          alignment=1, fontName='Helvetica-Bold', spaceAfter=6)
            styles = getSampleStyleSheet()
            story.append(Paragraph('Le Schéma — Rapport Transport', title_style))
            story.append(Paragraph(f'Abonnés: {len(students)} | Bus: {len(buses)}', styles['Normal']))
            story.append(Spacer(1, 8*mm))

            data = [['Code', 'Élève', 'Classe', 'Parent', 'Téléphone']]
            for s in students:
                data.append([s.code or '', f'{s.first_name} {s.last_name}',
                              s.class_name or '', s.parent_name or '', s.parent_phone or ''])
            t = Table(data, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0044cc')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f0f4ff')]),
                ('PADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(t)
            doc.build(story)
            self.status_lbl.setText(f'✅ Export transport: {path}')
        except Exception as e:
            QMessageBox.critical(self, 'Erreur', str(e))

    def _export_analytics_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.chart import BarChart, Reference
            from models.database import Payment
            from themes.style import MONTHS

            path, _ = QFileDialog.getSaveFileName(self, 'Enregistrer', f'analytiques_{datetime.now().year}.xlsx', 'Excel (*.xlsx)')
            if not path: return

            year = int(self.year_combo.currentText())
            payments = self.session.query(Payment).filter(Payment.year == year).all()

            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = 'Analyse Mensuelle'

            header_fill = PatternFill(start_color='0044CC', end_color='0044CC', fill_type='solid')
            headers = ['Mois', 'Mensualités', 'Assurance', 'Transport', 'Total']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill; cell.font = Font(color='FFFFFF', bold=True)

            for row, m in enumerate(MONTHS, 2):
                monthly = sum(p.amount for p in payments if p.month == m and p.payment_type == 'monthly')
                insurance = sum(p.amount for p in payments if p.month == m and p.payment_type == 'insurance')
                transport = sum(p.amount for p in payments if p.month == m and p.payment_type == 'transport')
                total = monthly + insurance + transport
                ws.cell(row=row, column=1, value=m)
                ws.cell(row=row, column=2, value=monthly)
                ws.cell(row=row, column=3, value=insurance)
                ws.cell(row=row, column=4, value=transport)
                ws.cell(row=row, column=5, value=total)

            chart = BarChart()
            chart.title = f'Revenus {year}'
            chart.y_axis.title = 'MAD'
            data_ref = Reference(ws, min_col=2, max_col=5, min_row=1, max_row=13)
            cats = Reference(ws, min_col=1, min_row=2, max_row=13)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, 'G2')

            wb.save(path)
            self.status_lbl.setText(f'✅ Analytics Excel: {path}')
        except Exception as e:
            QMessageBox.critical(self, 'Erreur', str(e))

    def _backup_db(self):
        import shutil
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        db_path = os.path.join(BASE_DIR, 'database', 'school.db')
        backup_dir = os.path.join(BASE_DIR, 'backups')
        os.makedirs(backup_dir, exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        dest = os.path.join(backup_dir, f'school_backup_{ts}.db')
        shutil.copy2(db_path, dest)
        self.status_lbl.setText(f'✅ Sauvegarde: {dest}')
        QMessageBox.information(self, 'Sauvegarde', f'Base de données sauvegardée:\n{dest}')
